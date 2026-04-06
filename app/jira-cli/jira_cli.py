#!/usr/bin/env python3
"""
jira-cli — Read and write Jira issues as structured Markdown.

Fetches the issue summary, description, comments, and all attached documents
(text, PDF, Word) and outputs a single clean structured document.
Can also update the issue description and add comments.

Usage (read):
  python jira_cli.py PROJECT-123
  python jira_cli.py PROJECT-123 --output report.md
  python jira_cli.py PROJECT-123 --no-attachments
  python jira_cli.py PROJECT-123 --comments-limit 10

Usage (write):
  python jira_cli.py PROJECT-123 --add-comment "My comment"
  python jira_cli.py PROJECT-123 --add-comment -            # read comment text from stdin
  python jira_cli.py PROJECT-123 --update-description "New description"
  python jira_cli.py PROJECT-123 --update-description -     # read description from stdin
  python jira_cli.py PROJECT-123 --attach-file path/to/file.feature  # upload a file as an attachment

Required environment variables (or .env file):
  JIRA_URL        — Jira base URL, e.g. https://yourorg.atlassian.net
  JIRA_USER       — Jira account email
  JIRA_API_TOKEN  — Jira API token (create at https://id.atlassian.com/manage-profile/security/api-tokens)
"""

import argparse
import io
import os
import sys
import textwrap
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from jira import JIRA, JIRAError
from rich.console import Console
from rich.progress import Progress, SpinnerColumn, TextColumn

load_dotenv()

console = Console(stderr=True)


# ---------------------------------------------------------------------------
# Jira connection
# ---------------------------------------------------------------------------

def get_jira_client() -> JIRA:
    url = os.environ.get("JIRA_URL", "").rstrip("/")
    user = os.environ.get("JIRA_USER", "")
    token = os.environ.get("JIRA_API_TOKEN", "")

    missing = [name for name, val in [("JIRA_URL", url), ("JIRA_USER", user), ("JIRA_API_TOKEN", token)] if not val]
    if missing:
        console.print(
            f"[red]Error:[/red] Missing required environment variable(s): {', '.join(missing)}\n"
            "Copy .env.example to .env and fill in your credentials.",
            highlight=False,
        )
        sys.exit(1)

    try:
        return JIRA(server=url, basic_auth=(user, token))
    except JIRAError as exc:
        console.print(f"[red]Error connecting to Jira:[/red] {exc.text}", highlight=False)
        sys.exit(1)


# ---------------------------------------------------------------------------
# Attachment text extraction
# ---------------------------------------------------------------------------

def _extract_text_from_bytes(filename: str, data: bytes) -> str:
    """Return plain text extracted from attachment bytes. Returns empty string on failure."""
    name_lower = filename.lower()

    # Plain text variants
    if any(name_lower.endswith(ext) for ext in (".txt", ".md", ".rst", ".csv", ".json", ".yaml", ".yml", ".xml", ".log")):
        try:
            return data.decode("utf-8", errors="replace")
        except Exception:
            return ""

    # PDF
    if name_lower.endswith(".pdf"):
        try:
            from pdfminer.high_level import extract_text_to_fp
            from pdfminer.layout import LAParams

            buf = io.StringIO()
            extract_text_to_fp(io.BytesIO(data), buf, laparams=LAParams(), output_type="text", codec=None)
            return buf.getvalue()
        except ImportError:
            return "[pdfminer.six not installed — cannot extract PDF text]"
        except Exception as exc:
            return f"[PDF extraction error: {exc}]"

    # Word (.docx)
    if name_lower.endswith(".docx"):
        try:
            import docx  # python-docx

            doc = docx.Document(io.BytesIO(data))
            return "\n".join(para.text for para in doc.paragraphs)
        except ImportError:
            return "[python-docx not installed — cannot extract .docx text]"
        except Exception as exc:
            return f"[DOCX extraction error: {exc}]"

    # Excel (.xlsx) — list sheet names and first rows
    if name_lower.endswith(".xlsx"):
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"Sheet: {sheet_name}")
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 20:
                        lines.append("... (truncated at 20 rows)")
                        break
                    lines.append("\t".join("" if v is None else str(v) for v in row))
            return "\n".join(lines)
        except ImportError:
            return "[openpyxl not installed — cannot extract .xlsx text]"
        except Exception as exc:
            return f"[XLSX extraction error: {exc}]"

    # Unsupported binary
    return f"[Binary file — content not extracted (size: {len(data):,} bytes)]"


def read_attachment(jira: JIRA, attachment) -> str:
    """Download an attachment and return its extracted text content."""
    try:
        data = jira._session.get(attachment.content).content
        return _extract_text_from_bytes(attachment.filename, data)
    except Exception as exc:
        return f"[Download error: {exc}]"


# ---------------------------------------------------------------------------
# Formatting helpers
# ---------------------------------------------------------------------------

def _format_date(date_str: Optional[str]) -> str:
    if not date_str:
        return "—"
    try:
        dt = datetime.fromisoformat(date_str.replace("Z", "+00:00"))
        return dt.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    except Exception:
        return date_str


def _indent(text: str, spaces: int = 2) -> str:
    prefix = " " * spaces
    return "\n".join(prefix + line for line in text.splitlines())


def _hr(char: str = "-", width: int = 72) -> str:
    return char * width


# ---------------------------------------------------------------------------
# Write operations
# ---------------------------------------------------------------------------

def _read_text_arg(value: str) -> str:
    """Return value as-is, or read from stdin if value is '-'."""
    if value == "-":
        return sys.stdin.read()
    return value


def add_comment(jira: JIRA, issue_key: str, body: str) -> None:
    """Add a comment to a Jira issue."""
    try:
        jira.add_comment(issue_key, body)
        console.print(f"[green]Comment added[/green] to {issue_key}")
    except JIRAError as exc:
        console.print(f"[red]Error adding comment to {issue_key}:[/red] {exc.text}", highlight=False)
        sys.exit(1)


def update_description(jira: JIRA, issue_key: str, description: str) -> None:
    """Replace the description of a Jira issue."""
    try:
        issue = jira.issue(issue_key)
        issue.update(fields={"description": description})
        console.print(f"[green]Description updated[/green] for {issue_key}")
    except JIRAError as exc:
        console.print(f"[red]Error updating description for {issue_key}:[/red] {exc.text}", highlight=False)
        sys.exit(1)


def attach_file(jira: JIRA, issue_key: str, file_path: str) -> None:
    """Upload a local file as an attachment to a Jira issue."""
    path = Path(file_path)
    if not path.exists():
        console.print(f"[red]Error:[/red] File not found: {file_path}", highlight=False)
        sys.exit(1)
    try:
        with open(path, "rb") as fh:
            jira.add_attachment(issue=issue_key, attachment=fh, filename=path.name)
        console.print(f"[green]File attached[/green] to {issue_key}: {path.name}")
    except JIRAError as exc:
        console.print(f"[red]Error attaching file to {issue_key}:[/red] {exc.text}", highlight=False)
        sys.exit(1)


# ---------------------------------------------------------------------------
# Main formatter
# ---------------------------------------------------------------------------

def format_issue(jira: JIRA, issue_key: str, include_attachments: bool = True, comments_limit: int = 0) -> str:
    """Fetch issue and return a structured Markdown document."""

    console.print(f"[cyan]Fetching[/cyan] {issue_key} …")

    try:
        issue = jira.issue(issue_key, expand="renderedFields")
    except JIRAError as exc:
        console.print(f"[red]Error fetching issue {issue_key}:[/red] {exc.text}", highlight=False)
        sys.exit(1)

    fields = issue.fields
    lines: list[str] = []

    # ── Header ──────────────────────────────────────────────────────────────
    lines += [
        f"# {issue_key}: {fields.summary}",
        "",
        _hr("="),
        "",
    ]

    # ── Metadata table ───────────────────────────────────────────────────────
    def _field(label: str, value: str) -> str:
        return f"| **{label}** | {value} |"

    status = getattr(fields.status, "name", "—")
    issue_type = getattr(fields.issuetype, "name", "—")
    priority = getattr(fields.priority, "name", "—") if fields.priority else "—"
    assignee = getattr(fields.assignee, "displayName", "Unassigned") if fields.assignee else "Unassigned"
    reporter = getattr(fields.reporter, "displayName", "—") if fields.reporter else "—"
    created = _format_date(getattr(fields, "created", None))
    updated = _format_date(getattr(fields, "updated", None))
    labels = ", ".join(fields.labels) if fields.labels else "—"
    fix_versions = ", ".join(v.name for v in fields.fixVersions) if fields.fixVersions else "—"
    components = ", ".join(c.name for c in fields.components) if fields.components else "—"

    lines += [
        "## Metadata",
        "",
        "| Field | Value |",
        "|---|---|",
        _field("Issue Key", issue_key),
        _field("Type", issue_type),
        _field("Status", status),
        _field("Priority", priority),
        _field("Assignee", assignee),
        _field("Reporter", reporter),
        _field("Created", created),
        _field("Updated", updated),
        _field("Labels", labels),
        _field("Fix Versions", fix_versions),
        _field("Components", components),
        "",
    ]

    # ── Description ──────────────────────────────────────────────────────────
    lines += ["## Description", ""]
    description = (fields.description or "").strip()
    if description:
        lines += [description, ""]
    else:
        lines += ["*No description provided.*", ""]

    # ── Attachments ──────────────────────────────────────────────────────────
    attachments = getattr(fields, "attachment", []) or []
    if attachments:
        lines += [f"## Attachments ({len(attachments)})", ""]

        if include_attachments:
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                console=console,
                transient=True,
            ) as progress:
                task = progress.add_task("Downloading attachments…", total=len(attachments))
                for att in attachments:
                    att_created = _format_date(getattr(att, "created", None))
                    att_author = getattr(att.author, "displayName", "—") if hasattr(att, "author") else "—"
                    lines += [
                        f"### {att.filename}",
                        "",
                        f"- **Size:** {int(att.size):,} bytes",
                        f"- **Uploaded by:** {att_author}",
                        f"- **Uploaded on:** {att_created}",
                        "",
                        "**Content:**",
                        "",
                        "```",
                    ]
                    content = read_attachment(jira, att)
                    # Truncate very large attachments
                    if len(content) > 10_000:
                        content = content[:10_000] + "\n\n... [truncated at 10 000 characters]"
                    lines += [content.strip(), "```", ""]
                    progress.advance(task)
        else:
            for att in attachments:
                att_created = _format_date(getattr(att, "created", None))
                att_author = getattr(att.author, "displayName", "—") if hasattr(att, "author") else "—"
                lines += [
                    f"- **{att.filename}** — {int(att.size):,} bytes — uploaded by {att_author} on {att_created}"
                ]
            lines.append("")

    # ── Comments ─────────────────────────────────────────────────────────────
    all_comments = jira.comments(issue)
    if comments_limit > 0:
        displayed_comments = all_comments[-comments_limit:]
        omitted = len(all_comments) - len(displayed_comments)
    else:
        displayed_comments = all_comments
        omitted = 0

    lines += [f"## Comments ({len(all_comments)})", ""]

    if omitted:
        lines += [f"*Showing last {comments_limit} of {len(all_comments)} comments.*", ""]

    if displayed_comments:
        for i, comment in enumerate(displayed_comments, start=1):
            author = getattr(comment.author, "displayName", "—") if comment.author else "—"
            created = _format_date(getattr(comment, "created", None))
            updated_c = _format_date(getattr(comment, "updated", None))
            updated_note = f" (edited {updated_c})" if updated_c != created else ""
            lines += [
                f"### Comment {i} — {author} on {created}{updated_note}",
                "",
                (comment.body or "").strip(),
                "",
            ]
    else:
        lines += ["*No comments.*", ""]

    # ── Footer ───────────────────────────────────────────────────────────────
    lines += [
        _hr("="),
        "",
        f"*Generated by jira-cli on {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}*",
    ]

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        prog="jira-cli",
        description="Fetch a Jira issue (description, comments, attachments) and output a structured Markdown document.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=textwrap.dedent("""\
            Examples (read):
              python jira_cli.py PROJECT-123
              python jira_cli.py PROJECT-123 --output report.md
              python jira_cli.py PROJECT-123 --no-attachments
              python jira_cli.py PROJECT-123 --comments-limit 5

            Examples (write):
              python jira_cli.py PROJECT-123 --add-comment "Reviewed and approved."
              python jira_cli.py PROJECT-123 --add-comment -
              python jira_cli.py PROJECT-123 --update-description "New description text"
              python jira_cli.py PROJECT-123 --update-description -
              python jira_cli.py PROJECT-123 --attach-file tests/scrum_12.feature
        """),
    )
    parser.add_argument("issue_key", help="Jira issue key, e.g. PROJECT-123")
    parser.add_argument(
        "-o", "--output",
        metavar="FILE",
        help="Write output to FILE instead of stdout",
    )
    parser.add_argument(
        "--no-attachments",
        action="store_true",
        help="Skip downloading attachment content (list filenames only)",
    )
    parser.add_argument(
        "--comments-limit",
        metavar="N",
        type=int,
        default=0,
        help="Only show the last N comments (0 = show all, default)",
    )
    parser.add_argument(
        "--add-comment",
        metavar="TEXT",
        help="Add a comment to the issue. Use '-' to read text from stdin.",
    )
    parser.add_argument(
        "--update-description",
        metavar="TEXT",
        help="Replace the issue description. Use '-' to read text from stdin.",
    )
    parser.add_argument(
        "--attach-file",
        metavar="PATH",
        help="Upload a local file as an attachment to the issue.",
    )
    parser.add_argument(
        "--env-file",
        metavar="PATH",
        default=".env",
        help="Path to .env file (default: .env)",
    )

    args = parser.parse_args()

    # Load env from specified file
    env_path = Path(args.env_file)
    if env_path.exists():
        load_dotenv(env_path, override=True)

    jira = get_jira_client()
    issue_key = args.issue_key.upper()

    # ── Write operations (mutually exclusive of the read/output flow) ────────
    write_requested = (
        args.add_comment is not None
        or args.update_description is not None
        or args.attach_file is not None
    )
    if write_requested:
        if args.update_description is not None:
            update_description(jira, issue_key, _read_text_arg(args.update_description))
        if args.add_comment is not None:
            add_comment(jira, issue_key, _read_text_arg(args.add_comment))
        if args.attach_file is not None:
            attach_file(jira, issue_key, args.attach_file)
        return

    # ── Read / format operation ───────────────────────────────────────────────
    result = format_issue(
        jira,
        issue_key,
        include_attachments=not args.no_attachments,
        comments_limit=args.comments_limit,
    )

    if args.output:
        out_path = Path(args.output)
        out_path.write_text(result, encoding="utf-8")
        console.print(f"[green]Saved[/green] → {out_path}")
    else:
        # Write structured output to stdout (progress/errors go to stderr)
        sys.stdout.write(result)
        sys.stdout.write("\n")


if __name__ == "__main__":
    main()
